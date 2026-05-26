"""
NaOH 三效降膜蒸发计算软件
NaOH Triple-Effect Falling-Film Evaporation Calculator
Chinese-language desktop GUI — exportable to Excel, packagable with PyInstaller.
"""

import sys
import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from naoh_evaporation import naoh_evaporation

# ── Fonts & colours ────────────────────────────────────────────────────────────
_FONT_FACE = "Microsoft YaHei"

FN  = (_FONT_FACE, 10)
FNB = (_FONT_FACE, 10, "bold")
FT  = (_FONT_FACE, 13, "bold")
FS  = (_FONT_FACE, 9)

BG        = "#F0F4F8"
BG_CARD   = "#FFFFFF"
HDR_BG    = "#1B4F8A"
HDR_FG    = "#FFFFFF"
ACCENT    = "#2471A3"
GREEN     = "#1E8449"
LABEL_FG  = "#2C3E50"
VALUE_FG  = "#1A5276"

# ── Chinese stream names ────────────────────────────────────────────────────────
_STREAM_CN = [
    "进料 (32% NaOH)",
    "生蒸汽",
    "V1 二次蒸汽 (EV101→EV201)",
    "V1 冷凝水",
    "V2 二次蒸汽 (EV201→EV301)",
    "V2 冷凝水",
    "V3 二次蒸汽 (EV301→冷凝器)",
    "L3 出 EV301",
    "L3a → E201 冷侧",
    "L3b → E202 冷侧",
    "L3 预热后 → EV201",
    "L2 出 EV201",
    "L2a → E101 冷侧",
    "L2b → E102 冷侧",
    "L2 预热后 → EV101",
    "L1 出 EV101 (50% NaOH)",
    "L1 经 E101 后 → E201",
    "50% NaOH 产品",
    "EV101 冷凝水 → E102",
    "冷凝水经 E102 后 → E202",
    "冷凝水产品",
]


# ── Main application ────────────────────────────────────────────────────────────
class NaOHCalculator(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("NaOH 三效降膜蒸发计算软件 v1.0")
        self.configure(bg=BG)
        self.minsize(1150, 720)
        self._results = None
        self._build_ui()

    # ── UI construction ─────────────────────────────────────────────────────────
    def _build_ui(self):
        self._build_header()
        self._build_inputs()
        self._build_buttons()
        self._build_notebook()

    def _build_header(self):
        hf = tk.Frame(self, bg=HDR_BG)
        hf.pack(fill="x")
        tk.Label(hf, text="NaOH 三效逆流降膜蒸发计算软件",
                 font=FT, bg=HDR_BG, fg=HDR_FG, pady=10).pack()

    def _build_inputs(self):
        outer = tk.LabelFrame(self, text="  输入参数  ", font=FNB,
                              bg=BG, fg=LABEL_FG, bd=1, relief="groove",
                              padx=12, pady=8)
        outer.pack(fill="x", padx=15, pady=(10, 4))

        self._vars = {}

        # (label_zh, key, default, unit)
        left_params = [
            ("进料流量 (F₀)",        "F0",           "10000.0", "kg/h"),
            ("进料温度 (T_feed)",     "T_feed",       "75.0",    "°C"),
            ("生蒸汽压力 (P_s)",      "P_s",          "10.0",    "bar"),
            ("冷端温差 ΔT_app",       "DT_APP_PH34",  "5.0",     "°C"),
        ]
        right_params = [
            ("一效压力 P₁ (EV101)",   "P1",              "2.092", "bar"),
            ("二效压力 P₂ (EV201)",   "P2",              "0.490", "bar"),
            ("三效压力 P₃ (EV301)",   "P3",              "0.100", "bar"),
            ("一效进料过热度 ΔT_sh",  "dT_superheat_1",  "6.0",   "°C"),
        ]

        for col_base, params in enumerate([left_params, right_params]):
            frame = tk.Frame(outer, bg=BG)
            frame.grid(row=0, column=col_base, padx=(0 if col_base == 0 else 30, 0), sticky="nw")
            for row, (label, key, default, unit) in enumerate(params):
                tk.Label(frame, text=label, font=FN, bg=BG, fg=LABEL_FG,
                         anchor="e", width=24).grid(row=row, column=0, pady=3, sticky="e")
                var = tk.StringVar(value=default)
                self._vars[key] = var
                tk.Entry(frame, textvariable=var, font=FN,
                         width=10, justify="right").grid(row=row, column=1, padx=6, pady=3)
                tk.Label(frame, text=unit, font=FS, bg=BG,
                         fg="#777777").grid(row=row, column=2, sticky="w")

    def _build_buttons(self):
        bf = tk.Frame(self, bg=BG)
        bf.pack(pady=6)

        def _btn(parent, text, color, hover, cmd):
            b = tk.Button(parent, text=text, font=FNB,
                          bg=color, fg="white",
                          activebackground=hover, activeforeground="white",
                          relief="flat", padx=18, pady=6, cursor="hand2",
                          command=cmd)
            b.pack(side="left", padx=8)
            return b

        _btn(bf, "  计  算  ",     ACCENT, "#1A5276", self._calculate)
        _btn(bf, "  导出 Excel  ", GREEN,  "#155232", self._export_excel)

        self._status = tk.StringVar(value='请输入参数后点击"计算"')
        tk.Label(bf, textvariable=self._status, font=FS,
                 bg=BG, fg="#555555").pack(side="left", padx=16)

    def _build_notebook(self):
        nb_frame = tk.Frame(self, bg=BG)
        nb_frame.pack(fill="both", expand=True, padx=15, pady=(0, 10))

        self._nb = ttk.Notebook(nb_frame)
        self._nb.pack(fill="both", expand=True)

        self._tab_summary = tk.Frame(self._nb, bg=BG_CARD)
        self._tab_streams  = tk.Frame(self._nb, bg=BG_CARD)
        self._tab_equip    = tk.Frame(self._nb, bg=BG_CARD)

        self._nb.add(self._tab_summary, text="  结果汇总  ")
        self._nb.add(self._tab_streams,  text="  流股数据  ")
        self._nb.add(self._tab_equip,    text="  设备汇总  ")

        self._init_stream_tree()
        self._init_equip_tree()
        self._summary_inner = None   # built after first calculation

    # ── Tree initialisers ───────────────────────────────────────────────────────
    def _init_stream_tree(self):
        cols    = ("no", "name", "frm", "to", "T", "P", "F", "h")
        headers = ("编号", "流股名称", "来源", "去向",
                   "温度 (°C)", "压力 (bar)", "流量 (kg/h)", "焓 (kcal/kg)")
        widths  = (45, 230, 80, 80, 95, 95, 105, 115)

        f = self._tab_streams
        sb = tk.Scrollbar(f, orient="vertical")
        sb.pack(side="right", fill="y")
        self._stree = ttk.Treeview(f, columns=cols, show="headings",
                                   yscrollcommand=sb.set, height=22)
        self._stree.pack(side="left", fill="both", expand=True)
        sb.config(command=self._stree.yview)

        for col, hdr, w in zip(cols, headers, widths):
            self._stree.heading(col, text=hdr)
            anchor = "w" if col in ("name", "frm", "to") else "center"
            self._stree.column(col, width=w, anchor=anchor, minwidth=w)

        self._stree.tag_configure("odd",  background="#F4F8FC")
        self._stree.tag_configure("even", background="#FFFFFF")

    def _init_equip_tree(self):
        cols    = ("unit", "hi", "ho", "ci", "co", "lmtd", "duty")
        headers = ("设备", "热侧进口 (°C)", "热侧出口 (°C)",
                   "冷侧进口 (°C)", "冷侧出口 (°C)", "LMTD (°C)", "热负荷 (Mcal/h)")
        widths  = (75, 120, 120, 120, 120, 105, 130)

        f = self._tab_equip
        self._etree = ttk.Treeview(f, columns=cols, show="headings", height=9)
        self._etree.pack(fill="both", expand=True, padx=4, pady=4)

        for col, hdr, w in zip(cols, headers, widths):
            self._etree.heading(col, text=hdr)
            self._etree.column(col, width=w, anchor="center", minwidth=w)

        self._etree.tag_configure("evap",    background="#EBF5FB")
        self._etree.tag_configure("preheat", background="#EAFAF1")

    # ── Calculation ──────────────────────────────────────────────────────────────
    def _calculate(self):
        try:
            p = {k: float(v.get()) for k, v in self._vars.items()}
        except ValueError as e:
            messagebox.showerror("输入错误", f"请检查输入参数：{e}")
            return

        self._status.set("计算中…")
        self.update_idletasks()

        try:
            r = naoh_evaporation(
                F0=p["F0"], T_feed=p["T_feed"], P_s=p["P_s"],
                P1=p["P1"], P2=p["P2"], P3=p["P3"],
                dT_superheat_1=p["dT_superheat_1"],
                DT_APP_PH34=p["DT_APP_PH34"],
                metric="full_results",
            )
        except Exception as exc:
            messagebox.showerror("计算错误", str(exc))
            self._status.set("计算出错")
            return

        if r is None:
            messagebox.showwarning("无可行解",
                "当前参数无可行解，请检查：\n"
                "• 各效压力约束 P₃ < P₂ < P₁\n"
                "• LMTD ≥ 9 °C（各蒸发器）\n"
                "• 进料温度低于生蒸汽温度")
            self._status.set("无可行解")
            return

        self._results = r
        self._fill_stream_tree(r["streams"])
        self._fill_equip_tree(r)
        self._fill_summary(r)
        self._status.set(
            f"计算完成  |  蒸汽消耗率 = {r['steam_per_tonne_naoh']:.2f} kg/t NaOH  "
            f"|  生蒸汽用量 D = {r['D']:.1f} kg/h"
        )

    # ── Fill stream tree ─────────────────────────────────────────────────────────
    def _fill_stream_tree(self, streams):
        for row in self._stree.get_children():
            self._stree.delete(row)
        for i, s in enumerate(streams):
            tag = "odd" if i % 2 == 0 else "even"
            name = _STREAM_CN[i] if i < len(_STREAM_CN) else s["name"]
            self._stree.insert("", "end", tags=(tag,), values=(
                s["no"], name, s["from"], s["to"],
                f"{s['T_C']:.1f}", f"{s['P_bar']:.3f}",
                f"{s['F_kg_h']:.0f}", f"{s['h_kcal_kg']:.2f}",
            ))

    # ── Fill equipment tree ──────────────────────────────────────────────────────
    def _fill_equip_tree(self, r):
        for row in self._etree.get_children():
            self._etree.delete(row)
        rows = [
            ("EV101", r["T_s"],   r["T_s"],         r["T1"], r["T1"],
             r["lmtd"]["EV101"], r["duty_mcal_h"]["EV101"], "evap"),
            ("EV201", r["T1_pure"], r["T1_pure"],   r["T2"], r["T2"],
             r["lmtd"]["EV201"], r["duty_mcal_h"]["EV201"], "evap"),
            ("EV301", r["T2_pure"], r["T2_pure"],   r["T3"], r["T3"],
             r["lmtd"]["EV301"], r["duty_mcal_h"]["EV301"], "evap"),
            ("E101", r["T1"],    r["T_mid"],         r["T2"], r["T_F1"],
             r["lmtd"]["E101"], r["duty_mcal_h"]["E101"], "preheat"),
            ("E102", r["T_s"],   r["T_mid"],         r["T2"], r["T_F1"],
             r["lmtd"]["E102"], r["duty_mcal_h"]["E102"], "preheat"),
            ("E201", r["T_mid"], r["T_hot_out34"],   r["T3"], r["T_F2"],
             r["lmtd"]["E201"], r["duty_mcal_h"]["E201"], "preheat"),
            ("E202", r["T_mid"], r["T_hot_out34"],   r["T3"], r["T_F2"],
             r["lmtd"]["E202"], r["duty_mcal_h"]["E202"], "preheat"),
        ]
        for unit, hi, ho, ci, co, lmtd, duty, tag in rows:
            self._etree.insert("", "end", tags=(tag,), values=(
                unit,
                f"{hi:.1f}", f"{ho:.1f}", f"{ci:.1f}", f"{co:.1f}",
                f"{lmtd:.2f}", f"{duty:.3f}",
            ))

    # ── Fill summary tab ──────────────────────────────────────────────────────────
    def _fill_summary(self, r):
        if self._summary_inner:
            self._summary_inner.destroy()

        outer = tk.Frame(self._tab_summary, bg=BG_CARD)
        outer.pack(fill="both", expand=True, padx=20, pady=15)
        self._summary_inner = outer

        def make_card(parent, title, items, col, row_start=0):
            card = tk.LabelFrame(parent, text=f"  {title}  ", font=FNB,
                                 bg=BG_CARD, fg=LABEL_FG, bd=1, relief="groove",
                                 padx=10, pady=8)
            card.grid(row=row_start, column=col, padx=8, pady=6, sticky="nw")
            for i, (lbl, val) in enumerate(items):
                tk.Label(card, text=lbl + "：", font=FN,
                         bg=BG_CARD, fg=LABEL_FG, anchor="e",
                         width=22).grid(row=i, column=0, sticky="e", pady=2)
                tk.Label(card, text=val, font=FNB,
                         bg=BG_CARD, fg=VALUE_FG, anchor="w",
                         width=20).grid(row=i, column=1, sticky="w", padx=6, pady=2)

        make_card(outer, "主要结果", [
            ("生蒸汽用量 D",    f"{r['D']:.1f} kg/h"),
            ("蒸汽消耗率",      f"{r['steam_per_tonne_naoh']:.2f} kg/t NaOH"),
            ("总蒸发量",        f"{r['total_evaporation_kg_h']:.1f} kg/h"),
            ("产品流量 L₁",     f"{r['L1']:.1f} kg/h"),
        ], col=0)

        make_card(outer, "各效浓度", [
            ("二效 NaOH 浓度 x₂", f"{r['x2']:.2f} %"),
            ("三效 NaOH 浓度 x₃", f"{r['x3']:.2f} %"),
        ], col=1)

        make_card(outer, "蒸发量分布", [
            ("一效蒸发量 W₁", f"{r['W1']:.1f} kg/h"),
            ("二效蒸发量 W₂", f"{r['W2']:.1f} kg/h"),
            ("三效蒸发量 W₃", f"{r['W3']:.1f} kg/h"),
        ], col=1, row_start=1)

        make_card(outer, "温度汇总", [
            ("一效沸点 T₁",          f"{r['T1']:.1f} °C"),
            ("二效沸点 T₂",          f"{r['T2']:.1f} °C"),
            ("三效沸点 T₃",          f"{r['T3']:.1f} °C"),
            ("一效进料温度 T_F1",    f"{r['T_F1']:.1f} °C"),
            ("二效进料温度 T_F2",    f"{r['T_F2']:.1f} °C"),
            ("预热器中间温度 T_mid", f"{r['T_mid']:.1f} °C"),
            ("产品侧出口温度",        f"{r['T_hot_out34']:.1f} °C"),
            ("生蒸汽温度 T_s",       f"{r['T_s']:.1f} °C"),
        ], col=2)

    # ── Excel export ────────────────────────────────────────────────────────────
    def _export_excel(self):
        if self._results is None:
            messagebox.showinfo("提示", '请先点击【计算】后再导出。')
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile="NaOH蒸发计算结果.xlsx",
        )
        if not path:
            return
        try:
            _write_excel(path, self._results, self._vars)
            messagebox.showinfo("导出成功", f"结果已保存至：\n{path}")
        except Exception as exc:
            messagebox.showerror("导出失败", str(exc))


# ── Excel writing ──────────────────────────────────────────────────────────────
def _write_excel(path, r, input_vars):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise ImportError("导出功能需要 openpyxl，请运行: pip install openpyxl")

    def hdr_cell(c, bg="1B4F8A", fg="FFFFFF"):
        c.font = Font(name="微软雅黑", bold=True, color=fg)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="AAAAAA")
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def data_cell(c, align="center", bold=False, bg=None):
        c.font = Font(name="微软雅黑", bold=bold)
        c.alignment = Alignment(horizontal=align, vertical="center")
        thin = Side(style="thin", color="DDDDDD")
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "结果汇总"
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "NaOH 三效降膜蒸发计算结果"
    hdr_cell(t, bg="1B4F8A")
    ws.row_dimensions[1].height = 28

    ws["A3"] = "输入参数"
    hdr_cell(ws["A3"], bg="2471A3")
    ws["D3"] = "计算结果"
    hdr_cell(ws["D3"], bg="1E8449")

    in_rows = [
        ("进料流量 F₀",       f"{float(input_vars['F0'].get()):.0f} kg/h"),
        ("进料温度 T_feed",    f"{float(input_vars['T_feed'].get()):.1f} °C"),
        ("生蒸汽压力 P_s",     f"{float(input_vars['P_s'].get()):.1f} bar"),
        ("一效压力 P₁",        f"{float(input_vars['P1'].get()):.3f} bar"),
        ("二效压力 P₂",        f"{float(input_vars['P2'].get()):.3f} bar"),
        ("三效压力 P₃",        f"{float(input_vars['P3'].get()):.3f} bar"),
        ("一效过热度 ΔT_sh",   f"{float(input_vars['dT_superheat_1'].get()):.1f} °C"),
        ("冷端温差 ΔT_app",    f"{float(input_vars['DT_APP_PH34'].get()):.1f} °C"),
    ]
    out_rows = [
        ("生蒸汽用量 D",          f"{r['D']:.1f} kg/h"),
        ("蒸汽消耗率",             f"{r['steam_per_tonne_naoh']:.2f} kg/t NaOH"),
        ("总蒸发量",               f"{r['total_evaporation_kg_h']:.1f} kg/h"),
        ("产品流量 L₁",            f"{r['L1']:.1f} kg/h"),
        ("二效浓度 x₂",            f"{r['x2']:.2f} %"),
        ("三效浓度 x₃",            f"{r['x3']:.2f} %"),
        ("一效沸点 T₁",            f"{r['T1']:.1f} °C"),
        ("二效沸点 T₂",            f"{r['T2']:.1f} °C"),
        ("三效沸点 T₃",            f"{r['T3']:.1f} °C"),
        ("一效进料温度 T_F1",      f"{r['T_F1']:.1f} °C"),
        ("二效进料温度 T_F2",      f"{r['T_F2']:.1f} °C"),
        ("预热器中间温度 T_mid",   f"{r['T_mid']:.1f} °C"),
        ("产品侧出口温度",          f"{r['T_hot_out34']:.1f} °C"),
        ("一效蒸发量 W₁",          f"{r['W1']:.1f} kg/h"),
        ("二效蒸发量 W₂",          f"{r['W2']:.1f} kg/h"),
        ("三效蒸发量 W₃",          f"{r['W3']:.1f} kg/h"),
    ]
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 20

    for i, (lbl, val) in enumerate(in_rows, start=4):
        c_lbl = ws.cell(row=i, column=1, value=lbl)
        c_val = ws.cell(row=i, column=2, value=val)
        data_cell(c_lbl, align="right")
        data_cell(c_val, align="left", bold=True)

    for i, (lbl, val) in enumerate(out_rows, start=4):
        c_lbl = ws.cell(row=i, column=4, value=lbl)
        c_val = ws.cell(row=i, column=5, value=val)
        data_cell(c_lbl, align="right")
        data_cell(c_val, align="left", bold=True, bg="EBF5FB")

    # ── Sheet 2: Stream table ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("流股数据")
    s_headers = ["编号", "流股名称", "来源", "去向",
                 "温度 (°C)", "压力 (bar)", "流量 (kg/h)", "焓 (kcal/kg)"]
    s_widths   = [8, 30, 12, 12, 14, 14, 14, 16]
    for ci, (hdr, w) in enumerate(zip(s_headers, s_widths), start=1):
        c = ws2.cell(row=1, column=ci, value=hdr)
        hdr_cell(c)
        ws2.column_dimensions[c.column_letter].width = w
    ws2.row_dimensions[1].height = 22
    ws2.freeze_panes = "A2"

    fills = ["EBF5FB", "FFFFFF"]
    for i, s in enumerate(r["streams"]):
        row = i + 2
        bg = fills[i % 2]
        name = _STREAM_CN[i] if i < len(_STREAM_CN) else s["name"]
        row_vals = [s["no"], name, s["from"], s["to"],
                    round(s["T_C"], 1), round(s["P_bar"], 3),
                    round(s["F_kg_h"]), round(s["h_kcal_kg"], 2)]
        for ci, val in enumerate(row_vals, start=1):
            c = ws2.cell(row=row, column=ci, value=val)
            align = "left" if ci in (2, 3, 4) else "center"
            data_cell(c, align=align, bg=bg)

    # ── Sheet 3: Equipment ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("设备汇总")
    e_headers = ["设备", "热侧进口 (°C)", "热侧出口 (°C)",
                 "冷侧进口 (°C)", "冷侧出口 (°C)", "LMTD (°C)", "热负荷 (Mcal/h)"]
    e_widths   = [10, 16, 16, 16, 16, 14, 18]
    for ci, (hdr, w) in enumerate(zip(e_headers, e_widths), start=1):
        c = ws3.cell(row=1, column=ci, value=hdr)
        hdr_cell(c)
        ws3.column_dimensions[c.column_letter].width = w
    ws3.row_dimensions[1].height = 22

    equip_rows = [
        ("EV101", r["T_s"],   r["T_s"],         r["T1"], r["T1"],
         r["lmtd"]["EV101"], r["duty_mcal_h"]["EV101"], "EBF5FB"),
        ("EV201", r["T1_pure"], r["T1_pure"],   r["T2"], r["T2"],
         r["lmtd"]["EV201"], r["duty_mcal_h"]["EV201"], "EBF5FB"),
        ("EV301", r["T2_pure"], r["T2_pure"],   r["T3"], r["T3"],
         r["lmtd"]["EV301"], r["duty_mcal_h"]["EV301"], "EBF5FB"),
        ("E101", r["T1"],    r["T_mid"],         r["T2"], r["T_F1"],
         r["lmtd"]["E101"], r["duty_mcal_h"]["E101"], "EAFAF1"),
        ("E102", r["T_s"],   r["T_mid"],         r["T2"], r["T_F1"],
         r["lmtd"]["E102"], r["duty_mcal_h"]["E102"], "EAFAF1"),
        ("E201", r["T_mid"], r["T_hot_out34"],   r["T3"], r["T_F2"],
         r["lmtd"]["E201"], r["duty_mcal_h"]["E201"], "EAFAF1"),
        ("E202", r["T_mid"], r["T_hot_out34"],   r["T3"], r["T_F2"],
         r["lmtd"]["E202"], r["duty_mcal_h"]["E202"], "EAFAF1"),
    ]
    for i, (unit, hi, ho, ci_t, co, lmtd, duty, bg) in enumerate(equip_rows, start=2):
        for col_i, val in enumerate([unit,
                                     round(hi, 1), round(ho, 1),
                                     round(ci_t, 1), round(co, 1),
                                     round(lmtd, 2), round(duty, 3)], start=1):
            c = ws3.cell(row=i, column=col_i, value=val)
            data_cell(c, align="center", bg=bg)

    wb.save(path)


# ── Entry point ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = NaOHCalculator()
    app.mainloop()
